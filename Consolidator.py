import os
import sys
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook, Workbook
import shutil

class ExcelConsolidator(FileSystemEventHandler):
    def __init__(self, folder_to_watch, processed_folder, not_applicable_folder, master_file):
        self.folder_to_watch = folder_to_watch
        self.processed_folder = processed_folder
        self.not_applicable_folder = not_applicable_folder
        self.master_file = master_file

    def on_created(self, event):
        if event.is_directory:
            return

        file_path = event.src_path
        if file_path.endswith(('.xls', '.xlsx', '.xlsm')):
            self.process_excel(file_path)
            self.move_file(file_path, self.processed_folder)
        else:
            self.move_file(file_path, self.not_applicable_folder)

    def process_excel(self, file_path):
        if os.path.exists(self.master_file):
            wb_master = load_workbook(self.master_file)
        else:
            wb_master = Workbook()
            wb_master.remove(wb_master.active)

        wb_new = load_workbook(file_path)

        for sheet_name in wb_new.sheetnames:
            ws_new = wb_new[sheet_name]
            ws_master = wb_master.create_sheet(sheet_name)
            for row in ws_new.iter_rows():
                for cell in row:
                    ws_master.cell(row=cell.row, column=cell.column, value=cell.value)

        wb_master.save(self.master_file)

    def move_file(self, file_path, destination_folder):
        os.makedirs(destination_folder, exist_ok=True)
        shutil.move(file_path, os.path.join(destination_folder, os.path.basename(file_path)))

if __name__ == "__main__":
    folder_to_watch = input("Enter the path to the folder you want to watch: ")
    processed_folder = os.path.join(folder_to_watch, "processed")
    not_applicable_folder = os.path.join(folder_to_watch, "not_applicable")
    master_folder = os.path.join(folder_to_watch, "master_workbook")
    os.makedirs(master_folder, exist_ok=True)
    master_file = os.path.join(master_folder, "master_file.xlsx")

    consolidator = ExcelConsolidator(folder_to_watch, processed_folder, not_applicable_folder, master_file)
    observer = Observer()
    observer.schedule(consolidator, folder_to_watch, recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
